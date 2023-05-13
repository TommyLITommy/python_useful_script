import os
import re
import numpy as np
from openpyxl import Workbook
from openpyxl.chart import (LineChart, ScatterChart, Reference)
from openpyxl.utils import get_column_letter
from openpyxl.styles import *

LOG_FILE_EXTENSION = ".log"
LOG_RESPONDER_PREFIX = "GB_New_R"

def get_all_files_from_dir(dir):
    filename_list = []
    for file_dir in os.listdir(dir):
        if not os.path.isdir(file_dir):
            if not file_dir.endswith(LOG_FILE_EXTENSION):
                continue
                
            if not file_dir.startswith("Grass_New_R"):
                continue
                
            filename_list.append(file_dir)
            
    # filename_list.sort(key=lambda file:int(file.split('_')[-1]))
    filename_list.sort(key=lambda file:int(file.split('_')[-1].replace('.log', '')))
    return filename_list

item_list = [
    'None',
    "File Name",
    "Ranging Count",
    "PER",
    "Min",
    "Max",
    "Mean",
    "Max-Mean",
    "Mean-Min",
    "STD",
]

def tidy_sheet(sheet):
    # 设置一个字典用于保存列宽数据
    dims = {}

    # 遍历表格数据，获取自适应列宽数据
    for row in sheet.rows:
        for cell in row:
            if cell.value:
                # 遍历整个表格，把该列所有的单元格文本进行长度对比，找出最长的单元格
                # 在对比单元格文本时需要将中文字符识别为1.7个长度，英文字符识别为1个，这里只需要将文本长度直接加上中文字符数量即可
                # re.findall('([\u4e00-\u9fa5])', cell.value)能够识别大部分中文字符
                cell_len = 0.7*len(re.findall('([\u4e00-\u9fa5])', str(cell.value))) + len(str(cell.value))
                dims[cell.column] = max((dims.get(cell.column, 0), cell_len))            
    for col, value in dims.items():
        # 设置列宽，get_column_letter用于获取数字列号对应的字母列号，最后值+2是用来调整最终效果的
        sheet.column_dimensions[get_column_letter(col)].width = value + 2

    for row in sheet.rows:
        for cell in row:
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = Border(left=Side(style='thin'), bottom=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'))
            #cell.alignment = Alignment(horizontal='center', vertical='center')
            # 

def set_line_background(sheet, row, count):
    bg_fill = PatternFill('solid', fgColor="3CB371")
    for col in range(count):
        cell = sheet.cell(row, col + 1)
        cell.fill = bg_fill

row = 2

def parse_log_file(cur_sheet, file):
    global row
    file_name = os.path.basename(file)
    file_name_prefix = os.path.splitext(file_name)[0]
    print("file_name_prefix:", file_name_prefix)

    raw_distance_list = []
    rx_fail_list = []
    ranging_count = 0
    d_max = 0
    d_min = 0
    d_mean = 0
    d_std = 0

    with open(file, errors='ignore') as file_ob:
        line  = file_ob.readline()
        while line:
            raw_distance_re = re.search("Raw Distance (\d+)", line)
            if raw_distance_re:
                raw_distance_list.append(int(raw_distance_re.group(1)))

            rx_fail_re = re.search("UWB RX fail  0x(\d+)", line)
            if rx_fail_re:
                rx_fail_list.append(int(rx_fail_re.group(1), 16))

            line = file_ob.readline()

        ranging_count = len(raw_distance_list) + len(rx_fail_list)
        d_min = min(raw_distance_list)
        d_max = max(raw_distance_list)
        d_mean = round(np.mean(raw_distance_list), 2)
        d_max_minus_mean = d_max - d_mean
        d_mean_minus_min = d_mean - d_min
        d_std = round(np.std(raw_distance_list), 2)
        d_per = round(len(rx_fail_list) / ranging_count * 100, 2)

        column = 1
        e = cur_sheet.cell(row, column)
        e.value = file_name_prefix

        column += 1
        e = cur_sheet.cell(row, column)
        e.value = ranging_count

        column += 1
        e = cur_sheet.cell(row, column)
        e.value = d_per

        column += 1
        e = cur_sheet.cell(row, column)
        e.value = d_min

        column += 1
        e = cur_sheet.cell(row, column)
        e.value = d_max

        column += 1
        e = cur_sheet.cell(row, column)
        e.value = d_mean

        column += 1
        e = cur_sheet.cell(row, column)
        e.value = d_max_minus_mean

        column += 1
        e = cur_sheet.cell(row, column)
        e.value = d_mean_minus_min

        column += 1
        e = cur_sheet.cell(row, column)
        e.value = d_std
    row += 1

def parse_file_and_create_excel(root, files):
    print("root:", root)
    global row
    row = 2

    file_name_list = []
    for file in files:
        if not file.endswith(LOG_FILE_EXTENSION):
            continue

        if not file.startswith(LOG_RESPONDER_PREFIX):
            continue

        file_name_list.append(file)

    if len(file_name_list) == 0:
        return

    try:
        file_name_list.sort(key=lambda file:int(file.split('_')[-1].replace('.log', '')))
    except:
        pass
    
    wb = Workbook()
    cur_sheet = wb.create_sheet("PER_STD")
    for index, item in enumerate(item_list):
        if index == 0:
            continue
        e = cur_sheet.cell(1, index)
        if item == "PER":
            item += "(%)"
        e.value = item

    for file in file_name_list:
        full_path = os.path.join(root, file)
        parse_log_file(cur_sheet, full_path)

    del wb['Sheet']
    for sheet in wb.worksheets:
        tidy_sheet(sheet)
        set_line_background(sheet, 1, len(item_list) - 1)

    excel_path = os.path.join(root, "PER_STD.xlsx")
    wb.save(excel_path)

if __name__ == '__main__':
    for root, dirs, files in os.walk(".", topdown=False):
        parse_file_and_create_excel(root, files)



        

    


