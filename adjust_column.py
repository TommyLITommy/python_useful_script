import openpyxl
from openpyxl.utils import get_column_letter
import re
from openpyxl.styles import *

def tidy_sheet(sheet):
    # 设置一个字典用于保存列宽数据
    dims = {}

    # 遍历表格数据，获取自适应列宽数据
    for row in sheet.rows:
        for cell in row:
            if cell.value:
                # 遍历整个表格，把该列所有的单元格文本进行长度对比，找出最长的单元格
                # 在对比单元格文本时需要将中文字符识别为1.7个长度，英文字符识别为1个，这里只需要将文本长度直接加上中文字符数量即可
                # re.findall('([\u4e00-\u9fa5])', cell.value)能够识别大部分中文字符
                cell_len = 0.7*len(re.findall('([\u4e00-\u9fa5])', str(cell.value))) + len(str(cell.value))
                dims[cell.column] = max((dims.get(cell.column, 0), cell_len))            
    for col, value in dims.items():
        # 设置列宽，get_column_letter用于获取数字列号对应的字母列号，最后值+2是用来调整最终效果的
        sheet.column_dimensions[get_column_letter(col)].width = value + 2

    for row in sheet.rows:
        for cell in row:
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = Border(left=Side(style='thin'), bottom=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'))
            #cell.alignment = Alignment(horizontal='center', vertical='center')
            # 

def set_line_background(sheet, row, count):
    bg_fill = PatternFill('solid', fgColor="3CB371")
    for col in range(count):
        cell = sheet.cell(row, col + 1)
        cell.fill = bg_fill



if __name__ == '__main__':
    # 修改下述参数即可使用，Excel名称及Sheet名称即可
    outputXlName = 'per_log_parse.xlsx'
    inputSheetName = 'PER Summary'

    wb = openpyxl.load_workbook(outputXlName)
    ws = wb[inputSheetName]
    for sheet in wb.worksheets:
        print(sheet)
        tidy_sheet(sheet)

    wb.save(outputXlName) 
